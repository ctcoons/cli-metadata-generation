# excel_utils/fill_conditions.py
import sqlite3
import openpyxl
import os


def fill_conditions_in_worklist(
    project_id: int,
    db_path: str = "project.db",
    template_path: str = "worklist_template_blank.xlsx",
    output_dir: str = "excel_utils/outputs",
    output_filename: str | None = None,   # ← make it optional
) -> str:
    from datetime import datetime
    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"worklist_conditions_{timestamp}.xlsx"
    """
    Fills the 'Condition' and 'Well Type' columns in the User sheet
    using group_names from the SQLite database.
    Returns the full path to the saved file.
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database not found: {db_path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, output_filename)

    # Connect to DB
    groups = []
    project_name = None

    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()

        # Get groups
        cur.execute("""
            SELECT DISTINCT group_name
            FROM group_data
            WHERE project_id = ?
        """, (project_id,))
        groups = [row[0] for row in cur.fetchall()]

        # Get project name
        cur.execute("""
            SELECT project_name
            FROM project_data
            WHERE id = ?   -- assuming the PK column is called 'id' (not 'project_id')
        """, (project_id,))
        
        row = cur.fetchone()
        if row:
            project_name = row[0]
        else:
            project_name = "Unknown Project"  # or raise ValueError("Project not found")

    if not groups:
        raise ValueError("No groups/conditions found in database.")

    # Load and modify Excel
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["User"]

    # Well Type = column 31 (AE), Condition = column 32 (AF)
    start_row = 2
    for i, group in enumerate(groups):
        row = start_row + i
        sheet.cell(row=row, column=31).value = "Condition"
        sheet.cell(row=row, column=32).value = group

    sheet.cell(row=3, column=2).value = project_name

    wb.save(output_path)
    print(f"Saved filled worklist to: {output_path}")

    return output_path